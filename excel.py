import openpyxl

def create_excel_and_fill():
     #Crear excel y llenar 
    new_excel = xl.Workbook()
    sheet1=new_excel.active
    sheet1['A1'].value = "Names"
    sheet1['B1'].value = "Last Names"
    sheet1['A2'].value = "Sandro"
    sheet1['B2'].value = "Sanchez"
    new_excel.save('logs/result.xlsx')


def obtener_workbook():
#obtener un workbook
     registro_jugadores = openpyxl.load_workbook("logs/Sample.xlsx")
     print(type(registro_jugadores))

def obtener_hojas_excel():
#obtener hojas del excel
     registro_jugadores = openpyxl.load_workbook("logs/Sample.xlsx")
     print(registro_jugadores.get_sheet_names())

def ir_a_una_ubicacion_excel1():
#acceder a una ubicacion del excel
#manera 1
     registro_jugadores = openpyxl.load_workbook("logs/Sample.xlsx")
     ubicacion = registro_jugadores.get_sheet_by_name('Hoja1')
     print(ubicacion['a2'].value)

def ir_a_una_ubicacion_excel2():
#manera 2
     print(ubicacion.cell(row=5,column=2).value)

def obtener_rango_excel():
#acceder a un rango de celdas
     listar_varias_celdas=ubicacion['a1':'b3']
     for row in listar_varias_celdas:
          for celda in row:
               print(celda.value)

def imprimir_todaslasfilas_excel():
#acceder a todas las filas 
     all_rows=ubicacion.rows
     #print(all_rows[])



ir_a_una_ubicacion_excel1()

